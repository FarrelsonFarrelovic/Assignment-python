
import openpyxl as xl
import csv

file1 = 'Employee_data.csv'
file = 'Employee_data.xlsx'

# Load in the workbook
wb = xl.load_workbook(file)

sheet = wb['Sheet1']

old_domaine = 'helpinghands.cm'
new_domaine = 'handsinhand.org'

for i in range (2, sheet.max_row + 1):
      cell = sheet.cell(i, 2) #choosing the column i'm going to work with
      if old_domaine in cell.value : 
        Updated_email = (cell.value).replace(old_domaine, new_domaine) #replacing the old domaine by the new domaine
  
        sheet.cell(i, 2).value = Updated_email#
        
wb.save('Updated_cell.xlsx')
#wb.save('update_sheet.csv')



mylist = []

old_domaine = 'helpinghands.cm'
new_domaine = 'handsinhand.org'

with open('Employee_data.csv', 'r') as file: #r method to read the file
    myFile = csv.reader(file)# the file we just opened
    for row in myFile:
        mylist.append(row)

print("See the csv file:")
for i in range(0, len(mylist)):
    sh = str(mylist[i])
    if old_domaine in sh:
        sh = sh.replace(old_domaine, new_domaine)


    
    
#Now We are creating the file.csv

with open('Update.csv', 'w+') as file:
    myFile = csv.writer(file, delimiter = ' ')
    for i in range(len(mylist)):
        myFile.writerow(mylist[i])
